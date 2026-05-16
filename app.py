import os
import re
import csv
import io
import json
import uuid
import tempfile
from datetime import datetime
from flask import Flask, request, jsonify, send_file, render_template
from werkzeug.utils import secure_filename
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB limit
app.config["UPLOAD_FOLDER"] = tempfile.gettempdir()

ALLOWED_EXTENSIONS = {"pdf"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ── PDF parsing ────────────────────────────────────────────────────────────────

DATE_PATTERNS = [
    r"\b(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})\b",
    r"\b(\d{4}[\/\-\.]\d{1,2}[\/\-\.]\d{1,2})\b",
    r"\b(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{2,4})\b",
    r"\b((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},?\s+\d{2,4})\b",
    r"\b(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*)\b",
    r"\b((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2})\b",
]

AMOUNT_PATTERN = re.compile(r"-?\$?[\d,]+\.\d{2}")


def extract_tables_from_pdf(path):
    """Try structured table extraction first, fall back to text parsing."""
    transactions = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                rows = parse_table(table)
                transactions.extend(rows)

        if not transactions:
            for page in pdf.pages:
                text = page.extract_text() or ""
                rows = parse_text_lines(text)
                transactions.extend(rows)

    return transactions


def parse_table(table):
    if not table or len(table) < 2:
        return []

    header = [str(c).lower().strip() if c else "" for c in table[0]]
    date_col = find_col(header, ["date", "transaction date", "posted", "value date"])
    desc_col = find_col(header, ["description", "details", "merchant", "narrative", "particulars", "memo", "payee"])
    debit_col = find_col(header, ["debit", "withdrawal", "dr", "charge", "amount out"])
    credit_col = find_col(header, ["credit", "deposit", "cr", "amount in"])
    amount_col = find_col(header, ["amount", "transaction amount"])
    balance_col = find_col(header, ["balance", "running balance"])

    rows = []
    for row in table[1:]:
        if not row or all(not c for c in row):
            continue
        date = safe_get(row, date_col)
        description = safe_get(row, desc_col)
        debit = safe_get(row, debit_col)
        credit = safe_get(row, credit_col)
        amount = safe_get(row, amount_col)
        balance = safe_get(row, balance_col)

        if not date and not description:
            continue

        # Resolve amount
        if not amount:
            if debit and parse_number(debit) != 0:
                amount = f"-{clean_amount(debit)}"
            elif credit and parse_number(credit) != 0:
                amount = clean_amount(credit)

        rows.append({
            "date": normalize_date(date),
            "description": clean_text(description),
            "debit": clean_amount(debit),
            "credit": clean_amount(credit),
            "amount": clean_amount(amount),
            "balance": clean_amount(balance),
        })
    return rows


def parse_text_lines(text):
    """Heuristic line-by-line parser for unstructured PDF text."""
    rows = []
    lines = text.splitlines()
    for line in lines:
        line = line.strip()
        if not line:
            continue

        date_match = None
        for pat in DATE_PATTERNS:
            m = re.search(pat, line, re.IGNORECASE)
            if m:
                date_match = m
                break

        amounts = AMOUNT_PATTERN.findall(line)
        if not date_match or not amounts:
            continue

        date_str = date_match.group(1)
        # Remove date and amounts from line to get description
        desc = line
        desc = re.sub(re.escape(date_str), "", desc, count=1)
        for amt in amounts:
            desc = desc.replace(amt, "", 1)
        desc = re.sub(r"\s{2,}", " ", desc).strip()

        amount = amounts[-1] if amounts else ""
        balance = ""
        if len(amounts) >= 2:
            balance = amounts[-1]
            amount = amounts[-2]

        rows.append({
            "date": normalize_date(date_str),
            "description": desc,
            "debit": "",
            "credit": "",
            "amount": clean_amount(amount),
            "balance": clean_amount(balance),
        })
    return rows


# ── Helpers ────────────────────────────────────────────────────────────────────

def find_col(header, candidates):
    for c in candidates:
        for i, h in enumerate(header):
            if c in h:
                return i
    return None


def safe_get(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return str(row[idx] or "").strip()


def clean_text(s):
    if not s:
        return ""
    return re.sub(r"\s+", " ", s).strip()


def clean_amount(s):
    if not s:
        return ""
    s = re.sub(r"[^\d.\-,]", "", s)
    s = s.replace(",", "")
    try:
        val = float(s)
        return f"{val:.2f}"
    except ValueError:
        return s


def parse_number(s):
    try:
        return float(clean_amount(s))
    except (ValueError, TypeError):
        return 0


def normalize_date(s):
    if not s:
        return ""
    s = s.strip()
    formats = [
        "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y",
        "%d %b %Y", "%b %d, %Y", "%B %d, %Y", "%d %B %Y",
        "%d %b", "%b %d",
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(s, fmt)
            if dt.year == 1900:
                dt = dt.replace(year=datetime.now().year)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


# ── Export ─────────────────────────────────────────────────────────────────────

def to_csv_bytes(rows):
    if not rows:
        return b""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["date", "description", "amount", "debit", "credit", "balance"])
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def to_excel_bytes(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="EBF0F8")
    border = Border(
        bottom=Side(style="thin", color="D0D7E3"),
    )

    headers = ["Date", "Description", "Amount", "Debit", "Credit", "Balance"]
    col_widths = [14, 48, 14, 14, 14, 16]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 22

    for row_idx, row in enumerate(rows, 2):
        values = [
            row.get("date", ""),
            row.get("description", ""),
            row.get("amount", ""),
            row.get("debit", ""),
            row.get("credit", ""),
            row.get("balance", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 2))
            if row_idx % 2 == 0:
                cell.fill = alt_fill

        # Colour amount column red for negatives
        amt_cell = ws.cell(row=row_idx, column=3)
        try:
            if float(row.get("amount", "0") or "0") < 0:
                amt_cell.font = Font(color="C0392B")
            else:
                amt_cell.font = Font(color="1E8449")
        except ValueError:
            pass

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/parse", methods=["POST"])
def parse():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400
    if not allowed_file(file.filename):
        return jsonify({"error": "Only PDF files are supported"}), 400

    filename = secure_filename(file.filename)
    tmp_path = os.path.join(app.config["UPLOAD_FOLDER"], f"{uuid.uuid4()}_{filename}")
    try:
        file.save(tmp_path)
        rows = extract_tables_from_pdf(tmp_path)
        # Remove completely empty rows
        rows = [r for r in rows if any(r.values())]
        return jsonify({"rows": rows, "count": len(rows)})
    except Exception as e:
        return jsonify({"error": f"Failed to parse PDF: {str(e)}"}), 500
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


@app.route("/download", methods=["POST"])
def download():
    data = request.get_json()
    rows = data.get("rows", [])
    fmt = data.get("format", "csv").lower()

    if not rows:
        return jsonify({"error": "No data to export"}), 400

    if fmt == "csv":
        content = to_csv_bytes(rows)
        return send_file(
            io.BytesIO(content),
            mimetype="text/csv",
            as_attachment=True,
            download_name="bank_statement.csv",
        )
    elif fmt == "excel":
        content = to_excel_bytes(rows)
        return send_file(
            io.BytesIO(content),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="bank_statement.xlsx",
        )
    else:
        return jsonify({"error": "Unknown format"}), 400


if __name__ == "__main__":
    app.run(debug=True, port=5000)
